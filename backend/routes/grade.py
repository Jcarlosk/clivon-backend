import csv
import io
import json
from datetime import datetime
from itertools import zip_longest
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from ..core.database import get_conn, get_cursor, DEFAULT_SUBJECTS
from ..core.auth import get_current_teacher
from ..core.omr_engine import process_omr

router = APIRouter(tags=["Correção e Gabaritos"])

# ── Modelos ───────────────────────────────────────────────────────────────────

class AnswerKeyPayload(BaseModel):
    subject: str
    class_id: str  # O frontend manda o nome da turma (ex: "2º Ano A")
    total_questions: int
    answers: list[str]
    bimester: int = 1  # Valor padrão adicionado para a nova estrutura

class ScanPayload(BaseModel):
    student_name: str
    subject: str
    class_id: str
    image_base64: Optional[str] = None

class ConfirmPayload(BaseModel):
    result_id: int
    student_name: str
    answers_final: list[str]

class ClearHistoryPayload(BaseModel):
    subject: str

# ── Gabaritos ─────────────────────────────────────────────────────────────────

@router.post("/save_answer_key", status_code=201)
def save_answer_key(payload: AnswerKeyPayload, teacher=Depends(get_current_teacher)):
    if len(payload.answers) != payload.total_questions:
        raise HTTPException(400, "O número de respostas diverge do total.")

    conn = get_conn()
    cur = get_cursor(conn)

    # ISOLAMENTO MULTI-ESCOLA: Sempre validar o school_id
    cur.execute(
        "SELECT id FROM answer_keys WHERE teacher_id=%s AND school_id=%s AND subject=%s AND class_name=%s",
        (teacher["id"], teacher["school_id"], payload.subject, payload.class_id)
    )
    existing = cur.fetchone()

    if existing:
        cur.execute(
            "UPDATE answer_keys SET answers=%s, total_questions=%s, bimester=%s WHERE id=%s",
            (json.dumps(payload.answers), payload.total_questions, payload.bimester, existing["id"])
        )
        key_id = existing["id"]
    else:
        cur.execute(
            """INSERT INTO answer_keys (school_id, teacher_id, subject, class_name, total_questions, answers, bimester)
               VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id""",
            (teacher["school_id"], teacher["id"], payload.subject, payload.class_id, 
             payload.total_questions, json.dumps(payload.answers), payload.bimester)
        )
        key_id = cur.fetchone()["id"]

    conn.commit()
    cur.close()
    conn.close()
    return {"id": key_id, "message": "Gabarito guardado com sucesso."}

# ── Scan ──────────────────────────────────────────────────────────────────────

@router.post("/scan")
def scan(payload: ScanPayload, teacher=Depends(get_current_teacher)):
    conn = get_conn()
    cur = get_cursor(conn)

    # ISOLAMENTO MULTI-ESCOLA
    cur.execute(
        "SELECT * FROM answer_keys WHERE teacher_id=%s AND school_id=%s AND subject=%s AND class_name=%s",
        (teacher["id"], teacher["school_id"], payload.subject, payload.class_id)
    )
    key_row = cur.fetchone()

    if not key_row:
        # Fallback: se não achar para a turma específica, pega o último gabarito da matéria
        cur.execute(
            "SELECT * FROM answer_keys WHERE teacher_id=%s AND school_id=%s AND subject=%s ORDER BY id DESC LIMIT 1",
            (teacher["id"], teacher["school_id"], payload.subject)
        )
        key_row = cur.fetchone()

    if not key_row:
        cur.close()
        conn.close()
        raise HTTPException(404, f"Nenhum gabarito configurado para '{payload.subject}'.")

    correct_answers = key_row["answers"] if isinstance(key_row["answers"], list) else json.loads(key_row["answers"])
    total = key_row["total_questions"]
    bimester = key_row["bimester"]

    answers_read = correct_answers
    detections = []
    debug_b64 = None

    if payload.image_base64:
        omr = process_omr(payload.image_base64, total, correct_answers)

        print(f"[DEBUG] Total esperado: {total} | Respostas detetadas pelo OMR: {len(omr.answers) if omr.success else 'falhou'}")

        if omr.success:
            answers_read = omr.answers
            debug_b64 = omr.debug_image_b64
            detections = [
                {
                    "question": d.question,
                    "detected_answer": d.detected_answer,
                    "confidence": d.confidence,
                    "bubbles": [
                        {
                            "option": b["option"],
                            "filled": b["filled"],
                            "ratio": b["ratio"],
                            "cx": b.get("cx", 0.5),
                            "cy": b.get("cy", 0.5),
                            "r":  b.get("r", 0.03),
                        }
                        for b in d.bubbles
                    ],
                }
                for d in omr.detections
            ]

    correct_count = sum(
        a == c
        for a, c in zip_longest(answers_read, correct_answers, fillvalue=None)
    )
    wrong_count = total - correct_count
    score = round((correct_count / total) * 10, 2)

    cur.close()
    conn.close()

    return {
        "draft": True,
        "answer_key_id": key_row["id"],
        "student_name": payload.student_name,
        "subject": payload.subject,
        "bimester": bimester,
        "answers_read": answers_read,
        "correct_answers": correct_answers,
        "correct": correct_count,
        "wrong": wrong_count,
        "score": score,
        "detections": detections,
        "debug_image_b64": debug_b64,
    }

# ── Confirmar resultado ───────────────────────────────────────────────────────

@router.post("/confirm_result", status_code=201)
def confirm_result(payload: ConfirmPayload, teacher=Depends(get_current_teacher)):
    conn = get_conn()
    cur = get_cursor(conn)

    cur.execute(
        "SELECT answers, total_questions, subject, class_name, bimester FROM answer_keys WHERE id=%s AND school_id=%s",
        (payload.result_id, teacher["school_id"])
    )
    key_row = cur.fetchone()

    if not key_row:
        cur.close()
        conn.close()
        raise HTTPException(404, "Gabarito de referência não encontrado ou sem permissão de acesso.")

    correct_answers = key_row["answers"] if isinstance(key_row["answers"], list) else json.loads(key_row["answers"])
    total = key_row["total_questions"]

    correct_count = sum(
        a == c
        for a, c in zip_longest(payload.answers_final, correct_answers, fillvalue=None)
    )
    score = round((correct_count / total) * 10, 2)

    cur.execute(
        """INSERT INTO scan_results
           (school_id, teacher_id, answer_key_id, student_name, subject, class_name, bimester, answers_read, correct, wrong, score, confirmed)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, True) RETURNING id""",
        (
            teacher["school_id"], teacher["id"], payload.result_id, payload.student_name,
            key_row["subject"], key_row["class_name"], key_row["bimester"],
            json.dumps(payload.answers_final),
            correct_count, (total - correct_count), score
        )
    )

    new_id = cur.fetchone()["id"]
    conn.commit()
    cur.close()
    conn.close()

    return {"id": new_id, "score": score}

# ── Histórico de resultados ───────────────────────────────────────────────────

@router.get("/results")
def get_results(teacher=Depends(get_current_teacher)):
    """Retorna todos os resultados confirmados da escola e do professor logado."""
    conn = get_conn()
    cur = get_cursor(conn)

    cur.execute(
        """SELECT id, student_name, subject, class_name, correct, wrong, score, scanned_at, bimester
           FROM scan_results
           WHERE teacher_id=%s AND school_id=%s
           ORDER BY scanned_at DESC
           LIMIT 200""",
        (teacher["id"], teacher["school_id"])
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()

    return [
        {
            "id":           row["id"],
            "student_name": row["student_name"],
            "subject":      row["subject"],
            "class_id":     row["class_name"], # Frontend ainda mapeia como class_id nas tabelas visuais
            "bimester":     row["bimester"],
            "correct":      row["correct"],
            "wrong":        row["wrong"],
            "score":        float(row["score"]),
            "scanned_at":   row["scanned_at"].isoformat() if row["scanned_at"] else None,
        }
        for row in rows
    ]

# ── Limpar histórico ──────────────────────────────────────────────────────────

@router.delete("/results/clear")
def clear_results(payload: ClearHistoryPayload, teacher=Depends(get_current_teacher)):
    conn = get_conn()
    cur = get_cursor(conn)

    if payload.subject == "ALL":
        cur.execute("DELETE FROM scan_results WHERE teacher_id=%s AND school_id=%s", 
                    (teacher["id"], teacher["school_id"]))
    else:
        cur.execute(
            "DELETE FROM scan_results WHERE teacher_id=%s AND school_id=%s AND subject=%s",
            (teacher["id"], teacher["school_id"], payload.subject)
        )

    deleted = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()

    return {"deleted": deleted, "message": "Histórico apagado com sucesso."}

# ── Dashboard ─────────────────────────────────────────────────────────────────

@router.get("/dashboard_stats")
def dashboard_stats(teacher=Depends(get_current_teacher)):
    conn = get_conn()
    cur = get_cursor(conn)

    cur.execute("SELECT subject, class_name FROM answer_keys WHERE teacher_id=%s AND school_id=%s", 
                (teacher["id"], teacher["school_id"]))
    key_rows = cur.fetchall()
    configured_subjects = {row["subject"] for row in key_rows}

    class_by_subject = {}
    for row in key_rows:
        class_by_subject[row["subject"]] = row["class_name"]

    stats = []
    for subj in DEFAULT_SUBJECTS:
        cur.execute(
            "SELECT COUNT(*) as c, AVG(score) as a FROM scan_results WHERE teacher_id=%s AND school_id=%s AND subject=%s",
            (teacher["id"], teacher["school_id"], subj)
        )
        res = cur.fetchone()

        stats.append({
            "subject":        subj,
            "class_id":       class_by_subject.get(subj),
            "scans_done":     res["c"] or 0,
            "avg_score":      round(float(res["a"]), 2) if res["a"] else None,
            "has_answer_key": subj in configured_subjects,
        })

    cur.close()
    conn.close()
    return {"teacher_name": teacher["name"], "subjects": stats}

# ══════════════════════════════════════════════════════════════════════════════
# ── EXPORTAÇÃO ────────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

def _fetch_results_for_export(teacher_id: int, school_id: int, subject: str, cur):
    """Busca resultados da disciplina garantindo o isolamento da escola."""
    cur.execute(
        """SELECT student_name, subject, class_name, correct, wrong, score, scanned_at, bimester
           FROM scan_results
           WHERE teacher_id=%s AND school_id=%s AND subject=%s
           ORDER BY student_name ASC""",
        (teacher_id, school_id, subject)
    )
    return cur.fetchall()

# ── 1. Exportar Planilha de Notas (.xlsx) ─────────────────────────────────────

@router.get("/export/csv")
def export_csv(
    subject: str = Query(...),
    teacher=Depends(get_current_teacher)
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "Biblioteca openpyxl não instalada. Execute: pip install openpyxl")

    conn = get_conn()
    cur = get_cursor(conn)
    rows = _fetch_results_for_export(teacher["id"], teacher["school_id"], subject, cur)
    cur.close()
    conn.close()

    if not rows:
        raise HTTPException(404, f"Nenhum resultado encontrado para '{subject}'.")

    wb = Workbook()
    ws = wb.active
    ws.title = subject[:30]

    header_fill  = PatternFill("solid", fgColor="2D5BE3")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Planilha de Notas — {subject}"
    title_cell.font  = Font(bold=True, size=13, color="111827")
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Exportado em {datetime.now().strftime('%d/%m/%Y %H:%M')} · Professor: {teacher['name']}"
    sub_cell.font  = Font(size=9, color="6B7280")
    sub_cell.alignment = center_align
    ws.row_dimensions[2].height = 18

    headers = ["Aluno", "Disciplina", "Turma", "Acertos", "Erros", "Nota (0–10)", "Data"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
    ws.row_dimensions[4].height = 22

    green_fill = PatternFill("solid", fgColor="DCFCE7")
    amber_fill = PatternFill("solid", fgColor="FEF9C3")
    red_fill   = PatternFill("solid", fgColor="FEE2E2")

    for r_idx, row in enumerate(rows, 5):
        score = float(row["score"])
        score_fill = green_fill if score >= 7 else (amber_fill if score >= 5 else red_fill)
        date_str = row["scanned_at"].strftime("%d/%m/%Y %H:%M") if row["scanned_at"] else "—"

        values = [
            row["student_name"],
            row["subject"],
            row["class_name"] or "—",
            row["correct"],
            row["wrong"],
            score,
            date_str,
        ]
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border    = thin_border
            cell.alignment = center_align if c_idx > 1 else Alignment(vertical="center")
            if c_idx == 6:  
                cell.fill = score_fill
                cell.font = Font(bold=True)

    col_widths = [28, 16, 12, 10, 10, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w

    total_row = len(rows) + 5
    ws.cell(row=total_row, column=1, value="MÉDIA GERAL").font = Font(bold=True)
    avg_score = sum(float(r["score"]) for r in rows) / len(rows)
    avg_cell = ws.cell(row=total_row, column=6, value=round(avg_score, 2))
    avg_cell.font = Font(bold=True)
    avg_cell.alignment = center_align

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"notas_{subject.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

# ── 2. Exportar Provas Corrigidas (.pdf) ──────────────────────────────────────

@router.get("/export/pdf")
def export_pdf(
    subject: str = Query(...),
    teacher=Depends(get_current_teacher)
):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        raise HTTPException(500, "Biblioteca reportlab não instalada. Execute: pip install reportlab")

    conn = get_conn()
    cur = get_cursor(conn)
    rows = _fetch_results_for_export(teacher["id"], teacher["school_id"], subject, cur)

    cur.execute(
        "SELECT answers, total_questions, class_name FROM answer_keys WHERE teacher_id=%s AND school_id=%s AND subject=%s ORDER BY id DESC LIMIT 1",
        (teacher["id"], teacher["school_id"], subject)
    )
    key_row = cur.fetchone()
    cur.close()
    conn.close()

    if not rows:
        raise HTTPException(404, f"Nenhum resultado encontrado para '{subject}'.")

    correct_answers = []
    if key_row:
        correct_answers = key_row["answers"] if isinstance(key_row["answers"], list) else json.loads(key_row["answers"])

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    styles = getSampleStyleSheet()
    BLUE   = colors.HexColor("#2D5BE3")
    GRAY   = colors.HexColor("#6B7280")
    GREEN  = colors.HexColor("#16A34A")
    RED    = colors.HexColor("#DC2626")
    AMBER  = colors.HexColor("#D97706")
    LIGHT  = colors.HexColor("#F0F2F7")

    title_style = ParagraphStyle("title", fontSize=16, fontName="Helvetica-Bold", textColor=BLUE, spaceAfter=4)
    sub_style   = ParagraphStyle("sub",   fontSize=9,  fontName="Helvetica",      textColor=GRAY, spaceAfter=12)

    story = []

    story.append(Paragraph(f"Provas Corrigidas — {subject}", title_style))
    story.append(Paragraph(
        f"Professor: {teacher['name']}  ·  Exportado em {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  {len(rows)} aluno(s)",
        sub_style
    ))

    header = ["Aluno", "Turma", "Acertos", "Erros", "Nota", "Data"]
    data   = [header]

    for row in rows:
        score    = float(row["score"])
        date_str = row["scanned_at"].strftime("%d/%m/%Y") if row["scanned_at"] else "—"
        data.append([
            row["student_name"],
            row["class_name"] or "—",
            str(row["correct"]),
            str(row["wrong"]),
            f"{score:.1f}",
            date_str,
        ])

    col_widths = [7*cm, 2.5*cm, 2*cm, 2*cm, 2*cm, 2.5*cm]
    t = Table(data, colWidths=col_widths, repeatRows=1)

    ts = TableStyle([
        ("BACKGROUND",   (0,0), (-1,0),  BLUE),
        ("TEXTCOLOR",    (0,0), (-1,0),  colors.white),
        ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,0),  9),
        ("ALIGN",        (0,0), (-1,0),  "CENTER"),
        ("TOPPADDING",   (0,0), (-1,0),  8),
        ("BOTTOMPADDING",(0,0), (-1,0),  8),
        ("FONTNAME",     (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",     (0,1), (-1,-1), 9),
        ("ALIGN",        (1,1), (-1,-1), "CENTER"),
        ("ALIGN",        (0,1), (0,-1),  "LEFT"),
        ("TOPPADDING",   (0,1), (-1,-1), 6),
        ("BOTTOMPADDING",(0,1), (-1,-1), 6),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, LIGHT]),
        ("GRID",         (0,0), (-1,-1), 0.4, colors.HexColor("#E5E7EB")),
        ("LINEBELOW",    (0,0), (-1,0),  1,   BLUE),
    ])

    for i, row in enumerate(rows, 1):
        score = float(row["score"])
        note_color = GREEN if score >= 7 else (AMBER if score >= 5 else RED)
        ts.add("TEXTCOLOR",  (4, i), (4, i), note_color)
        ts.add("FONTNAME",   (4, i), (4, i), "Helvetica-Bold")

    t.setStyle(ts)
    story.append(t)
    story.append(Spacer(1, 0.6*cm))

    if correct_answers:
        story.append(Paragraph("Gabarito Oficial", ParagraphStyle(
            "h2", fontSize=12, fontName="Helvetica-Bold", textColor=BLUE, spaceBefore=8, spaceAfter=6
        )))

        per_row = 10
        gab_data = []
        for i in range(0, len(correct_answers), per_row):
            chunk = correct_answers[i:i+per_row]
            nums  = [str(j+1) for j in range(i, i+len(chunk))]
            gab_data.append(nums)
            gab_data.append(chunk)

        gab_table = Table(gab_data, colWidths=[1.6*cm]*per_row)
        gab_table.setStyle(TableStyle([
            ("ALIGN",       (0,0), (-1,-1), "CENTER"),
            ("FONTSIZE",    (0,0), (-1,-1), 9),
            ("FONTNAME",    (0,0), (-1,-1), "Helvetica"),
            ("FONTNAME",    (0,0), (-1,0),  "Helvetica-Bold"),
            ("TEXTCOLOR",   (0,0), (-1,0),  GRAY),
            ("FONTNAME",    (0,1), (-1,-1), "Helvetica-Bold"),
            ("TEXTCOLOR",   (0,1), (-1,-1), BLUE),
            ("TOPPADDING",  (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0),(-1,-1), 5),
            ("ROWBACKGROUNDS",(0,0),(-1,-1),[LIGHT, colors.white]),
            ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#E5E7EB")),
        ]))
        story.append(gab_table)

    doc.build(story)
    buf.seek(0)

    filename = f"provas_{subject.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

# ── 3. Exportar Gabarito em Branco (.pdf) ────────────────────────────────────

@router.get("/export/gabarito")
def export_gabarito(
    subject: str = Query(...),
    teacher=Depends(get_current_teacher)
):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
    except ImportError:
        raise HTTPException(500, "Biblioteca reportlab não instalada. Execute: pip install reportlab")

    conn = get_conn()
    cur = get_cursor(conn)
    cur.execute(
        "SELECT answers, total_questions, class_name FROM answer_keys WHERE teacher_id=%s AND school_id=%s AND subject=%s ORDER BY id DESC LIMIT 1",
        (teacher["id"], teacher["school_id"], subject)
    )
    key_row = cur.fetchone()
    cur.close()
    conn.close()

    if not key_row:
        raise HTTPException(404, f"Nenhum gabarito configurado para '{subject}'.")

    total_q = key_row["total_questions"]
    class_id = key_row["class_name"] or ""

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    BLUE  = colors.HexColor("#2D5BE3")
    GRAY  = colors.HexColor("#6B7280")
    LIGHT = colors.HexColor("#F0F2F7")
    WHITE = colors.white

    story = []

    story.append(Paragraph(f"Folha de Respostas — {subject}", ParagraphStyle(
        "title", fontSize=15, fontName="Helvetica-Bold", textColor=BLUE, spaceAfter=2
    )))
    story.append(Paragraph(
        f"Turma: {class_id}  ·  Total de questões: {total_q}  ·  Professor: {teacher['name']}",
        ParagraphStyle("sub", fontSize=9, fontName="Helvetica", textColor=GRAY, spaceAfter=14)
    ))

    info_data = [["Aluno: ___________________________________________", "Data: ___/___/______"]]
    info_table = Table(info_data, colWidths=[12*cm, 5*cm])
    info_table.setStyle(TableStyle([
        ("FONTNAME",  (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE",  (0,0), (-1,-1), 10),
        ("TOPPADDING",(0,0), (-1,-1), 0),
        ("BOTTOMPADDING",(0,0),(-1,-1), 0),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.5*cm))

    story.append(Paragraph(
        "Instruções: Preencha apenas UMA alternativa por questão. Não rasure.",
        ParagraphStyle("inst", fontSize=8, fontName="Helvetica", textColor=GRAY, spaceAfter=10)
    ))

    OPTIONS = ["A", "B", "C", "D", "E"]
    header_row = ["Nº"] + OPTIONS
    data = [header_row]

    for q in range(1, total_q + 1):
        row = [str(q)] + ["◯"] * len(OPTIONS)
        data.append(row)

    col_w = [1.2*cm] + [1.6*cm] * len(OPTIONS)
    grid = Table(data, colWidths=col_w, repeatRows=1)
    grid.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  BLUE),
        ("TEXTCOLOR",     (0,0), (-1,0),  WHITE),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0),  10),
        ("ALIGN",         (0,0), (-1,0),  "CENTER"),
        ("TOPPADDING",    (0,0), (-1,0),  7),
        ("BOTTOMPADDING", (0,0), (-1,0),  7),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 13),
        ("ALIGN",         (0,1), (-1,-1), "CENTER"),
        ("TEXTCOLOR",     (0,1), (0,-1),  GRAY),
        ("FONTSIZE",      (0,1), (0,-1),  9),
        ("FONTNAME",      (0,1), (0,-1),  "Helvetica-Bold"),
        ("TOPPADDING",    (0,1), (-1,-1), 5),
        ("BOTTOMPADDING", (0,1), (-1,-1), 5),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),  [WHITE, LIGHT]),
        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#E5E7EB")),
        ("LINEBELOW",     (0,0), (-1,0),  1,   BLUE),
        ("BOX",           (0,0), (-1,-1), 0.8, BLUE),
    ]))
    story.append(grid)

    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph(
        f"Clivon Edu · Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle("footer", fontSize=7, fontName="Helvetica", textColor=GRAY)
    ))

    doc.build(story)
    buf.seek(0)

    filename = f"gabarito_branco_{subject.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )